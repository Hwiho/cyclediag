"""Per-cycle rest voltage sampling after charge / discharge legs (canonical)."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from typing import Any

import numpy as np
import pandas as pd

__all__ = [
    "RestVoltageSample",
    "RestPeriodResult",
    "CycleRestVoltageResult",
    "parse_time_offsets",
    "sample_voltage_at_offsets",
    "extract_cycle_rest_voltages",
    "flatten_rest_voltage_rows",
    "build_rest_voltage_workbook_sheets",
    "write_rest_voltage_excel",
]


def _resolve_current_column(df: pd.DataFrame) -> str | None:
    """Pick a current column (AvgCurrent preferred)."""
    priority: list[str] = []
    fallback: list[str] = []
    for col in df.columns:
        clean = str(col).lower().replace(" ", "").replace("_", "")
        if clean in ("avgcurrent", "avgcurrentma", "avgcurrenta"):
            priority.append(col)
        elif clean in ("current", "curr", "i", "currenta", "currentma"):
            fallback.append(col)
    for col in priority + fallback:
        series = pd.to_numeric(df[col], errors="coerce").abs()
        valid = series[np.isfinite(series) & (series > 0)]
        if len(valid) >= 4:
            return col
    return None


@dataclass
class RestVoltageSample:
    """Voltage at one offset within a rest period."""

    after_leg: str
    offset_s: float
    voltage: float | None
    status: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class RestPeriodResult:
    """All samples for one rest block (after charge or discharge)."""

    after_leg: str
    rest_duration_s: float | None
    samples: list[RestVoltageSample]
    method: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "after_leg": self.after_leg,
            "rest_duration_s": self.rest_duration_s,
            "samples": [s.to_dict() for s in self.samples],
            "method": self.method,
        }


@dataclass
class CycleRestVoltageResult:
    """Rest voltage readings for one cycle."""

    cycle: int
    periods: list[RestPeriodResult]

    def to_dict(self) -> dict[str, Any]:
        return {
            "cycle": self.cycle,
            "periods": [p.to_dict() for p in self.periods],
        }


def parse_time_offsets(text: str) -> list[float]:
    """Parse comma-separated seconds, e.g. ``'0.1, 60, 300'``."""
    if not str(text).strip():
        return []
    out: list[float] = []
    for part in str(text).split(","):
        token = part.strip().lower()
        if not token:
            continue
        if token.endswith("m") and token[:-1].strip().replace(".", "", 1).isdigit():
            out.append(float(token[:-1]) * 60.0)
        elif token.endswith("min"):
            out.append(float(token[:-3].strip()) * 60.0)
        else:
            out.append(float(token))
    return sorted(set(out))


def _label_set(text: str) -> set[str]:
    return {p.strip().lower() for p in str(text).split(",") if p.strip()}


def _classify_step(
    step: str,
    *,
    charge_labels: set[str],
    discharge_labels: set[str],
    rest_labels: set[str],
) -> str:
    s = str(step).strip().lower()
    if s in charge_labels:
        return "charge"
    if s in discharge_labels:
        return "discharge"
    for r in rest_labels:
        if s == r or r in s:
            return "rest"
    return "other"


def _classify_with_current(
    step_kind: str,
    current: float | None,
    *,
    rest_current_max: float | None,
) -> str:
    if step_kind != "other":
        return step_kind
    if (
        rest_current_max is not None
        and current is not None
        and np.isfinite(current)
        and abs(current) <= rest_current_max
    ):
        return "rest"
    return "other"


def _segment_kinds(
    steps: pd.Series,
    currents: pd.Series | None,
    *,
    charge_labels: set[str],
    discharge_labels: set[str],
    rest_labels: set[str],
    rest_current_max: float | None,
) -> list[str]:
    kinds: list[str] = []
    for i, step in enumerate(steps.astype(str)):
        cur = None
        if currents is not None and i < len(currents):
            cur = pd.to_numeric(currents.iloc[i], errors="coerce")
            cur = float(cur) if np.isfinite(cur) else None
        base = _classify_step(
            step,
            charge_labels=charge_labels,
            discharge_labels=discharge_labels,
            rest_labels=rest_labels,
        )
        kinds.append(
            _classify_with_current(base, cur, rest_current_max=rest_current_max)
        )
    return kinds


def _rest_relative_time(
    rest_df: pd.DataFrame,
    t_col: str | None,
    step_t_col: str | None,
) -> np.ndarray:
    if step_t_col and step_t_col in rest_df.columns:
        st = pd.to_numeric(rest_df[step_t_col], errors="coerce").to_numpy(dtype=float)
        if np.isfinite(st).sum() >= max(2, len(rest_df) // 2):
            st0 = st[np.isfinite(st)][0]
            return st - st0
    if t_col and t_col in rest_df.columns:
        t = pd.to_numeric(rest_df[t_col], errors="coerce").to_numpy(dtype=float)
        if np.isfinite(t).sum() >= 2:
            t0 = t[np.isfinite(t)][0]
            return t - t0
    return np.arange(len(rest_df), dtype=float)


def sample_voltage_at_offsets(
    rest_df: pd.DataFrame,
    offsets_s: list[float],
    *,
    v_col: str,
    t_col: str | None = None,
    step_t_col: str | None = None,
) -> tuple[list[RestVoltageSample], float | None]:
    if rest_df is None or rest_df.empty or v_col not in rest_df.columns:
        samples = [
            RestVoltageSample(after_leg="", offset_s=o, voltage=None, status="empty")
            for o in offsets_s
        ]
        return samples, None

    t_rel = _rest_relative_time(rest_df, t_col, step_t_col)
    v = pd.to_numeric(rest_df[v_col], errors="coerce").to_numpy(dtype=float)
    valid = np.isfinite(t_rel) & np.isfinite(v)
    duration = float(np.nanmax(t_rel[valid])) if valid.any() else None

    samples: list[RestVoltageSample] = []
    if valid.sum() == 0:
        for o in offsets_s:
            samples.append(RestVoltageSample("", o, None, "no_data"))
        return samples, duration

    t_arr = t_rel[valid]
    v_arr = v[valid]
    order = np.argsort(t_arr)
    t_arr = t_arr[order]
    v_arr = v_arr[order]

    for offset in offsets_s:
        if offset < t_arr[0] - 1e-9:
            samples.append(RestVoltageSample("", offset, None, "before_start"))
        elif offset > t_arr[-1] + 1e-9:
            samples.append(RestVoltageSample("", offset, None, "beyond_rest"))
        elif valid.sum() == 1:
            v0 = float(v_arr[0])
            samples.append(
                RestVoltageSample("", offset, v0 if abs(offset - t_arr[0]) < 1e-9 else None, "single_point")
            )
        else:
            samples.append(
                RestVoltageSample("", offset, float(np.interp(offset, t_arr, v_arr)), "ok")
            )
    return samples, duration


def extract_cycle_rest_voltages(
    cycle_df: pd.DataFrame,
    *,
    v_col: str,
    st_col: str,
    charge_label: str = "charge",
    discharge_label: str = "discharge",
    rest_labels_text: str = "rest",
    offsets_s: list[float] | None = None,
    t_col: str | None = "TotalTime_sec",
    step_t_col: str | None = "StepTime_sec",
    rest_current_max: float | None = None,
) -> CycleRestVoltageResult:
    """Find rest periods after charge/discharge and sample voltage at offsets."""
    offsets = list(offsets_s or [0.1, 60.0])
    cycle_val = None
    if cycle_df is not None and not cycle_df.empty and "TotalCycle" in cycle_df.columns:
        cycle_val = cycle_df["TotalCycle"].iloc[0]

    empty = CycleRestVoltageResult(cycle=int(cycle_val or 0), periods=[])
    if cycle_df is None or cycle_df.empty or st_col not in cycle_df.columns:
        return empty
    if v_col not in cycle_df.columns:
        return empty

    df = cycle_df.reset_index(drop=True)
    charge_labels = _label_set(charge_label)
    discharge_labels = _label_set(discharge_label)
    rest_labels = _label_set(rest_labels_text)

    i_col = _resolve_current_column(df)
    currents = pd.to_numeric(df[i_col], errors="coerce") if i_col else None
    kinds = _segment_kinds(
        df[st_col],
        currents,
        charge_labels=charge_labels,
        discharge_labels=discharge_labels,
        rest_labels=rest_labels,
        rest_current_max=rest_current_max,
    )

    segments: list[tuple[str, int, int]] = []
    if kinds:
        start = 0
        for i in range(1, len(kinds)):
            if kinds[i] != kinds[i - 1]:
                segments.append((kinds[i - 1], start, i - 1))
                start = i
        segments.append((kinds[-1], start, len(kinds) - 1))

    periods: list[RestPeriodResult] = []
    for idx, (kind, s_idx, e_idx) in enumerate(segments):
        if kind != "rest":
            continue
        prev_kind = segments[idx - 1][0] if idx > 0 else None
        if prev_kind not in ("charge", "discharge"):
            continue

        rest_df = df.iloc[s_idx : e_idx + 1]
        samples, duration = sample_voltage_at_offsets(
            rest_df,
            offsets,
            v_col=v_col,
            t_col=t_col,
            step_t_col=step_t_col,
        )
        for s in samples:
            s.after_leg = prev_kind
        method = "steptype"
        if i_col and rest_labels == {"rest"}:
            method = "steptype+current" if rest_current_max is not None else "steptype"
        periods.append(
            RestPeriodResult(
                after_leg=prev_kind,
                rest_duration_s=duration,
                samples=samples,
                method=method,
            )
        )

    return CycleRestVoltageResult(cycle=int(cycle_val or 0), periods=periods)


def flatten_rest_voltage_rows(
    results: list[CycleRestVoltageResult],
    *,
    filepath: str = "",
    fname: str = "",
    display_cycle: int | str | None = None,
    offsets_s: list[float] | None = None,
) -> list[dict[str, Any]]:
    """Flatten per-cycle results into table rows."""
    offsets = offsets_s or []
    rows: list[dict[str, Any]] = []
    for res in results:
        disp = display_cycle if display_cycle is not None else res.cycle
        if not res.periods:
            row: dict[str, Any] = {
                "file": fname,
                "filepath": filepath,
                "cycle": res.cycle,
                "display_cycle": disp,
                "after_leg": "",
                "rest_duration_s": None,
                "method": "no_rest",
            }
            for o in offsets:
                row[f"v_{o:g}s"] = None
                row[f"status_{o:g}s"] = "no_rest"
            rows.append(row)
            continue

        for period in res.periods:
            row = {
                "file": fname,
                "filepath": filepath,
                "cycle": res.cycle,
                "display_cycle": disp,
                "after_leg": period.after_leg,
                "rest_duration_s": period.rest_duration_s,
                "method": period.method,
            }
            for sample in period.samples:
                key = f"v_{sample.offset_s:g}s"
                row[key] = sample.voltage
                row[f"status_{sample.offset_s:g}s"] = sample.status
            rows.append(row)
    return rows


def _offset_col(offset_s: float) -> str:
    return f"V_{offset_s:g}s"


def _leg_label(after_leg: str) -> str:
    if after_leg == "charge":
        return "EoC (after charge)"
    if after_leg == "discharge":
        return "EoD (after discharge)"
    return after_leg or ""


def build_rest_voltage_workbook_sheets(
    rows: list[dict[str, Any]],
    offsets_s: list[float],
    *,
    meta: dict[str, Any] | None = None,
) -> dict[str, pd.DataFrame]:
    """Build Excel sheets: Meta / All / EoC / EoD from flattened rest rows."""
    v_cols = [_offset_col(o) for o in offsets_s]
    base_cols = ["file", "cycle", "after_leg", "rest_label", "rest_duration_s", "method"]

    records: list[dict[str, Any]] = []
    for r in rows:
        leg = r.get("after_leg") or ""
        rec: dict[str, Any] = {
            "file": r.get("file", ""),
            "cycle": r.get("display_cycle", r.get("cycle")),
            "after_leg": leg,
            "rest_label": _leg_label(leg),
            "rest_duration_s": r.get("rest_duration_s"),
            "method": r.get("method", ""),
        }
        for o in offsets_s:
            key = f"v_{o:g}s"
            rec[_offset_col(o)] = r.get(key)
        records.append(rec)

    all_df = pd.DataFrame(records)
    if all_df.empty:
        all_df = pd.DataFrame(columns=base_cols + v_cols)
    else:
        all_df = all_df[base_cols + v_cols]

    eoc = all_df[all_df["after_leg"] == "charge"].drop(
        columns=["after_leg", "rest_label"], errors="ignore"
    ).reset_index(drop=True)
    eod = all_df[all_df["after_leg"] == "discharge"].drop(
        columns=["after_leg", "rest_label"], errors="ignore"
    ).reset_index(drop=True)

    meta_rows = [
        {"item": "sample_times_s", "value": ", ".join(f"{o:g}" for o in offsets_s)},
        {"item": "n_rows", "value": len(all_df)},
        {"item": "n_eoc_rows", "value": len(eoc)},
        {"item": "n_eod_rows", "value": len(eod)},
        {"item": "n_files", "value": int(all_df["file"].nunique()) if not all_df.empty else 0},
    ]
    if meta:
        for k, v in meta.items():
            meta_rows.append({"item": str(k), "value": v})
    meta_df = pd.DataFrame(meta_rows)

    return {
        "Meta": meta_df,
        "All": all_df,
        "EoC": eoc,
        "EoD": eod,
    }


def write_rest_voltage_excel(
    path: str,
    sheets: dict[str, pd.DataFrame],
) -> None:
    """Write rest-voltage workbook to ``path`` (.xlsx)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe = str(name)[:31] or "Sheet"
            df.to_excel(writer, sheet_name=safe, index=False)
            ws = writer.sheets[safe]
            header_fill = PatternFill("solid", fgColor="DCE6F1")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_w = max(
                    (len(str(c.value)) for c in col_cells if c.value is not None),
                    default=8,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_w + 2, 28)
